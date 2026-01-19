"""Spreadsheet (Excel) extraction using openpyxl."""

import io
import logging

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from app.services.file.extraction.models import ExtractedDocument, TextBlock

logger = logging.getLogger(__name__)

# Maximum dimensions to prevent processing extremely large spreadsheets
MAX_ROWS = 10000
MAX_COLS = 100


def _cell_value_to_str(cell: Cell) -> str:
    """Convert a cell value to string, handling various types."""
    if cell.value is None:
        return ""
    if isinstance(cell.value, (int, float)):
        # Format numbers nicely (no trailing .0 for integers)
        if isinstance(cell.value, float) and cell.value.is_integer():
            return str(int(cell.value))
        return str(cell.value)
    return str(cell.value)


def _format_row_as_markdown(cells: list[str], is_header: bool = False) -> str:
    """Format a row of cells as markdown table row."""
    row = "| " + " | ".join(cells) + " |"
    if is_header:
        # Add separator row after header
        separator = "| " + " | ".join(["---"] * len(cells)) + " |"
        return row + "\n" + separator
    return row


class SpreadsheetExtractor:
    """Extract text from Excel spreadsheets (.xlsx, .xls)."""

    def extract(self, content: bytes, file_name: str | None = None) -> ExtractedDocument:
        """
        Extract text from an Excel spreadsheet.

        Converts each sheet to a markdown table representation.
        The entire spreadsheet is returned as a single TextBlock to preserve
        context (not chunked).

        Args:
            content: Raw bytes of the Excel file
            file_name: Optional file name for metadata

        Returns:
            ExtractedDocument with all sheets as TextBlocks
        """
        try:
            workbook = load_workbook(
                filename=io.BytesIO(content),
                read_only=True,
                data_only=True,  # Get calculated values, not formulas
            )
        except Exception as e:
            logger.error(f"Failed to load spreadsheet: {e}")
            return ExtractedDocument(
                title=file_name,
                blocks=[],
                metadata={"source_type": "spreadsheet", "error": str(e)},
            )

        blocks: list[TextBlock] = []
        sheet_count = len(workbook.sheetnames)

        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            try:
                sheet = workbook[sheet_name]
                sheet_text = self._extract_sheet(sheet, sheet_name)

                if sheet_text.strip():
                    blocks.append(
                        TextBlock(
                            text=sheet_text,
                            location={
                                "type": "sheet",
                                "sheet_name": sheet_name,
                                "sheet_index": sheet_index,
                            },
                            heading_context=f"Sheet: {sheet_name}",
                        )
                    )
            except Exception as e:
                logger.warning(f"Failed to extract sheet '{sheet_name}': {e}")
                continue

        workbook.close()

        return ExtractedDocument(
            title=file_name,
            blocks=blocks,
            metadata={
                "source_type": "spreadsheet",
                "sheet_count": sheet_count,
            },
        )

    def _extract_sheet(self, sheet, sheet_name: str) -> str:
        """
        Extract text from a single worksheet as markdown table.

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet for context

        Returns:
            Markdown-formatted text representation of the sheet
        """
        lines: list[str] = []
        lines.append(f"## {sheet_name}\n")

        # Determine actual data bounds
        max_row = min(sheet.max_row or 0, MAX_ROWS)
        max_col = min(sheet.max_column or 0, MAX_COLS)

        if max_row == 0 or max_col == 0:
            return f"## {sheet_name}\n\n*Empty sheet*"

        # Collect all rows with data
        rows_data: list[list[str]] = []
        for row_idx in range(1, max_row + 1):
            row_cells = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_cells.append(_cell_value_to_str(cell))
            # Only include rows that have at least some content
            if any(cell.strip() for cell in row_cells):
                rows_data.append(row_cells)

        if not rows_data:
            return f"## {sheet_name}\n\n*Empty sheet*"

        # Normalize column count (pad shorter rows)
        max_cols_in_data = max(len(row) for row in rows_data)
        for row in rows_data:
            while len(row) < max_cols_in_data:
                row.append("")

        # Format as markdown table
        # Treat first row as header
        if len(rows_data) >= 1:
            lines.append(_format_row_as_markdown(rows_data[0], is_header=True))
            for row in rows_data[1:]:
                lines.append(_format_row_as_markdown(row))

        # Add row count info if truncated
        if max_row >= MAX_ROWS:
            lines.append(f"\n*Note: Showing first {MAX_ROWS} rows (truncated)*")

        return "\n".join(lines)
